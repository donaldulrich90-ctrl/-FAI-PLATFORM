"""Génération des exports Excel / CSV pour l'app finance."""
from __future__ import annotations

import csv
import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Styles partagés ─────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill(start_color="0A3D62", end_color="0A3D62", fill_type="solid")
_HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TOT_FILL  = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
_TOT_FONT  = Font(color="FFFFFF", bold=True)

_ALT_FILL  = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
_LBL_FONT  = Font(bold=True, color="0A3D62")

_CENTER = Alignment(horizontal="center")
_RIGHT  = Alignment(horizontal="right")


def _header_row(ws, headers: list[str], col_widths: list[int]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill  = _HDR_FILL
        c.font  = _HDR_FONT
        c.alignment = _HDR_ALIGN
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _alt_row(ws, row: int, n_cols: int) -> None:
    if row % 2 == 0:
        for col in range(1, n_cols + 1):
            ws.cell(row=row, column=col).fill = _ALT_FILL


def _totals_row(ws, row: int, values: list, n_cols: int) -> None:
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.fill = _TOT_FILL
        c.font = _TOT_FONT
    for col in range(4, n_cols + 1):
        ws.cell(row=row, column=col).alignment = _RIGHT


# ── Export 1 : liste des rapports revendeurs ────────────────────────────────

def build_revendeur_reports_excel(qs) -> bytes:
    """Excel — liste des rapports journaliers revendeurs (1 feuille)."""
    from django.db.models import Sum

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapports revendeurs"

    headers    = ["Date", "Revendeur", "Préfixe", "Vendus", "Utilisés",
                  "Brut (XOF)", "Commission (XOF)", "Net FAI (XOF)"]
    col_widths = [13, 32, 12, 10, 10, 18, 18, 18]
    _header_row(ws, headers, col_widths)

    n = 0
    for row_idx, r in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1, value=r.report_date).alignment = _CENTER
        ws.cell(row=row_idx, column=2, value=r.revendeur.get_full_name() or r.revendeur.username)
        ws.cell(row=row_idx, column=3, value=r.prefix).alignment = _CENTER
        ws.cell(row=row_idx, column=4, value=r.tickets_sold_count).alignment = _RIGHT
        ws.cell(row=row_idx, column=5, value=r.tickets_used_count).alignment = _RIGHT
        ws.cell(row=row_idx, column=6, value=int(r.gross_xof)).alignment = _RIGHT
        ws.cell(row=row_idx, column=7, value=int(r.commission_xof)).alignment = _RIGHT
        ws.cell(row=row_idx, column=8, value=int(r.net_isp_xof)).alignment = _RIGHT
        _alt_row(ws, row_idx, len(headers))
        n += 1

    if n > 0:
        agg = qs.aggregate(
            gross=Sum("gross_xof"), comm=Sum("commission_xof"), net=Sum("net_isp_xof"),
            sold=Sum("tickets_sold_count"), used=Sum("tickets_used_count"),
        )
        _totals_row(ws, n + 2, [
            "TOTAUX", "", "",
            agg["sold"] or 0, agg["used"] or 0,
            int(agg["gross"] or 0), int(agg["comm"] or 0), int(agg["net"] or 0),
        ], len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Export 2 : rapport revendeur détaillé ───────────────────────────────────

_DURATIONS = {"3h": "3 heures", "1d": "24 heures", "1w": "7 jours", "30j": "30 jours"}
_STATUSES  = {"available": "Disponible", "used": "Utilisé", "expired": "Expiré"}


def build_revendeur_report_detail_excel(report) -> bytes:
    """Excel 2 feuilles — résumé financier + détail ticket par ticket."""
    wb = openpyxl.Workbook()

    # ── Feuille 1 : Résumé ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 24

    rows = [
        ("Revendeur",                   report.revendeur.get_full_name() or report.revendeur.username),
        ("Préfixe",                      report.prefix or "—"),
        ("Date du rapport",              report.report_date),
        ("Généré le",                    report.generated_at.strftime("%d/%m/%Y %H:%M") if report.generated_at else ""),
        ("", ""),
        ("Tickets vendus",               report.tickets_sold_count),
        ("Tickets utilisés",             report.tickets_used_count),
        ("", ""),
        ("Montant brut encaissé (XOF)",  int(report.gross_xof)),
        ("Commission revendeur (XOF)",   int(report.commission_xof)),
        ("Net FAI (XOF)",                int(report.net_isp_xof)),
    ]

    net_font = Font(bold=True, color="1B5E20", size=12)
    for row_idx, (label, value) in enumerate(rows, 2):
        c_lbl = ws1.cell(row=row_idx, column=1, value=label)
        c_val = ws1.cell(row=row_idx, column=2, value=value)
        if label:
            c_lbl.font = _LBL_FONT
        if label == "Net FAI (XOF)":
            c_val.font = net_font
        c_val.alignment = _RIGHT

    # ── Feuille 2 : Tickets ─────────────────────────────────────
    ws2 = wb.create_sheet("Tickets")
    headers    = ["Code", "Durée", "Site", "Prix (XOF)", "Commission (XOF)",
                  "Net FAI (XOF)", "Statut", "Hotspot", "Vendu le"]
    col_widths = [22, 12, 22, 13, 16, 14, 13, 10, 22]
    _header_row(ws2, headers, col_widths)

    for row_idx, t in enumerate(report.detail_json, 2):
        ws2.cell(row=row_idx, column=1, value=t.get("code", ""))
        ws2.cell(row=row_idx, column=2, value=_DURATIONS.get(t.get("duration", ""), t.get("duration", ""))).alignment = _CENTER
        ws2.cell(row=row_idx, column=3, value=t.get("site_name") or t.get("site", ""))
        ws2.cell(row=row_idx, column=4, value=int(Decimal(str(t.get("price_xof", 0))))).alignment = _RIGHT
        ws2.cell(row=row_idx, column=5, value=int(Decimal(str(t.get("commission_xof", 0))))).alignment = _RIGHT
        ws2.cell(row=row_idx, column=6, value=int(Decimal(str(t.get("net_isp_xof", 0))))).alignment = _RIGHT
        ws2.cell(row=row_idx, column=7, value=_STATUSES.get(t.get("status", ""), t.get("status", ""))).alignment = _CENTER
        ws2.cell(row=row_idx, column=8, value="Oui" if t.get("hotspot_synced") else "Non").alignment = _CENTER
        ws2.cell(row=row_idx, column=9, value=(t.get("sold_at") or "")[:16])
        _alt_row(ws2, row_idx, len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Export 3 : journal de caisse ────────────────────────────────────────────

def build_caisse_journal_csv(qs) -> str:
    """CSV du journal de caisse (CashJournalEntry) — séparateur ; pour Excel FR."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    writer.writerow(["Date", "Type", "Catégorie", "Description",
                     "Montant (XOF)", "Site", "Créé par", "Date saisie"])

    types = {"income": "Entrée", "expense": "Dépense"}
    for e in qs:
        writer.writerow([
            e.entry_date.strftime("%d/%m/%Y"),
            types.get(e.entry_type, e.entry_type),
            e.category or "",
            e.description,
            int(e.amount_xof),
            e.site.name if e.site else "",
            (e.created_by.get_full_name() or e.created_by.username) if e.created_by else "",
            e.created_at.strftime("%d/%m/%Y %H:%M"),
        ])

    return "﻿" + output.getvalue()  # BOM UTF-8 pour Excel Windows
